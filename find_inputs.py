import pandas as pd
from openpyxl import load_workbook

XLSX_PATH = "Base Pricing27.1_Pro_SMART_RCGV.xlsx"

wb = load_workbook(XLSX_PATH, data_only=True)
ws = wb["Input Sheet"]

print("Scanning for all input cells in Input Sheet:\n")
print("="*80)

# Scan more rows and columns
for row in range(1, 50):
    row_data = []
    for col in range(1, 20):  # Check first 20 columns
        cell = ws.cell(row, col)
        if cell.value and str(cell.value).strip():
            row_data.append(f"{cell.coordinate}: {cell.value}")
    if row_data:
        print(" | ".join(row_data))

wb.close()