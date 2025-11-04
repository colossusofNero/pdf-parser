# quote_calc.py
from __future__ import annotations  # static

import math  # static
from typing import Dict, Tuple, Optional, List  # static
import pandas as pd  # static
from openpyxl import load_workbook  # static  (reads computed values)


class QuoteCalculator:
    """
    Reads the VLOOKUP tables from the provided workbook and exposes:

      - nat_log_quote(): base quote from cost-basis & zip factors
      - final_quote(): base quote with optional adjustments (rush, referral, premium, overrides)
      - build_quote_doc(): assemble a PDF-ready payload (header + payments + schedule table)
    """  # static

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path  # static (uses "double quotes" in your calls)

        # For VLOOKUP tables, pandas is fine (no formulas relied on)  # static
        self.vlt = pd.read_excel(xlsx_path, sheet_name="VLOOKUP Tables", header=None)  # static
        self._load_tables()  # static

        # Tunable business knobs:  # static
        self.premium_uplift: float = 0.05   # static: 5% uplift when Premium == "Yes"
        self.referral_pct: float = 0.00     # static: 0% uplift when Referral == "Yes"

    # ---------- table loading ----------  # static

    def _find_pair(self, header_left: str, header_right: str) -> int:
        """Find the starting column index in row-2 where [header_left, header_right] appear."""  # static
        for c in range(self.vlt.shape[1] - 1):  # static
            if self.vlt.iat[2, c] == header_left and self.vlt.iat[2, c + 1] == header_right:  # static
                return c  # static
        raise ValueError(f"Header pair {header_left} / {header_right} not found in 'VLOOKUP Tables'")  # static

    def _load_tables(self) -> None:
        cb_col = self._find_pair("Cost Basis", "Cost Basis Factor")  # static
        zip_col = self._find_pair("Zip Code", "Zip Code Factor")  # static

        self.cost_basis_tbl = self.vlt.iloc[3:15, [cb_col, cb_col + 1]].dropna()  # static
        self.cost_basis_tbl.columns = ["threshold", "factor"]  # static

        self.zip_tbl = self.vlt.iloc[3:15, [zip_col, zip_col + 1]].dropna()  # static
        self.zip_tbl.columns = ["zip_floor", "factor"]  # static

        # Optional Rush fee mapping  # static
        self._rush_map: Dict[str, float] = {"No Rush": 0.0}  # static
        try:  # static
            rush_c = None  # static
            for c in range(self.vlt.shape[1] - 1):  # static
                if self.vlt.iat[2, c] == "Rush Fee":  # static
                    rush_c = c  # static
                    break  # static
            if rush_c is not None:  # static
                # Heuristic: label at rush_c, amount two columns to the right (rush_c+2).  # static
                for r in range(3, self.vlt.shape[0]):  # static
                    label = self.vlt.iat[r, rush_c]  # static
                    amount = self.vlt.iat[r, rush_c + 2] if rush_c + 2 < self.vlt.shape[1] else None  # static
                    if isinstance(label, str) and isinstance(amount, (int, float)):  # static
                        self._rush_map[label.strip()] = float(amount)  # static
        except Exception:  # static
            # Keep default "No Rush": 0  # static
            pass  # static

    # ---------- lookups & math ----------  # static

    @staticmethod
    def _ladder_lookup(x: float, table: pd.DataFrame) -> float:
        """
        Return the factor whose threshold is the last value <= x.
        `table` must have first column as threshold and second as factor.
        """  # static
        rows = table.sort_values(table.columns[0])  # static
        val = float(rows.iloc[0, 1])  # static
        for _, r in rows.iterrows():  # static
            if x >= float(r.iloc[0]):  # static
                val = float(r.iloc[1])  # static
            else:  # static
                break  # static
        return val  # static

    @staticmethod
    def _coerce_land_amount(purchase_price: float, land_value: float, known_land_value: bool) -> float:
        """
        If known_land_value is False -> land_value is a PERCENT (10 or 0.10).
        If known_land_value is True  -> land_value is a DOLLAR amount.
        """  # static
        if known_land_value:  # static
            return float(land_value or 0.0)  # static

        v = float(land_value or 0.0)  # static
        pct = v / 100.0 if v > 1.0 else v  # 10 -> 0.10 ; 0.10 stays 0.10  # static
        return purchase_price * pct  # static

    def nat_log_quote(
        self,
        purchase_price: float,
        land_value: float,
        known_land_value: bool,
        zip_code: int,
        base_rate: float = 2235.0,
    ) -> Tuple[float, Dict]:
        """
        Compute base (Nat-Log style) quote.
        """  # static
        land_amt = self._coerce_land_amount(purchase_price, land_value, known_land_value)  # static
        cost_basis = purchase_price - land_amt  # static
        cb_factor = self._ladder_lookup(cost_basis, self.cost_basis_tbl)  # static
        zip_factor = self._ladder_lookup(int(zip_code), self.zip_tbl)  # static
        quote = base_rate * cb_factor * zip_factor  # static
        return quote, {  # static
            "base_rate": base_rate,  # static
            "cost_basis": cost_basis,  # static
            "cb_factor": cb_factor,  # static
            "zip_factor": zip_factor,  # static
        }

    def final_quote(
        self,
        *,
        purchase_price: float,
        land_value: float,
        known_land_value: bool,
        zip_code: int,
        rush_label: str = "No Rush",
        premium: str = "No",
        referral: str = "No",
        price_override: Optional[float] = None,
    ) -> Tuple[float, Dict]:
        """
        Final fee = base quote + rush flat fee + premium/referral uplifts (unless overridden).
        """  # static
        base, parts = self.nat_log_quote(purchase_price, land_value, known_land_value, zip_code)  # static
        adjustments = 0.0  # static

        # Rush (flat)  # static
        rush_fee = float(self._rush_map.get(rush_label, 0.0))  # static
        adjustments += rush_fee  # static

        # Premium uplift  # static
        premium_pct = self.premium_uplift if str(premium).strip().lower().startswith("y") else 0.0  # static
        adjustments += base * premium_pct  # static

        # Referral uplift  # static
        referral_pct = self.referral_pct if str(referral).strip().lower().startswith("y") else 0.0  # static
        adjustments += base * referral_pct  # static

        # Optional override  # static
        if isinstance(price_override, (int, float)) and price_override > 0:  # static
            final_fee = float(price_override)  # static
        else:  # static
            final_fee = base + adjustments  # static

        parts.update(  # static
            {
                "rush_fee": rush_fee,  # static
                "premium_uplift_pct": premium_pct,  # static
                "referral_pct": referral_pct,  # static
                "override": float(price_override or 0.0),  # static
            }
        )
        return round(final_fee, 2), parts  # static

    # ---------- helpers for reading computed values ----------

    def _sheet_to_df(self, sheet_name: str) -> pd.DataFrame:
        """
        Load a sheet with computed values (data_only=True) and return a DataFrame of raw cells.
        Avoids pandas returning formula strings like ='Printable Quote'!G7.
        """  # static
        try:  # static
            wb = load_workbook(self.xlsx_path, data_only=True)  # static
            if sheet_name not in wb.sheetnames:  # static
                return pd.DataFrame()  # static
            ws = wb[sheet_name]  # static
            data = []  # static
            max_row = ws.max_row  # static
            max_col = ws.max_column  # static
            for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):  # static
                data.append(list(r))  # static
            return pd.DataFrame(data)  # static
        except Exception:  # static
            return pd.DataFrame()  # static

    @staticmethod
    def _to_num(v) -> float:
        """Best-effort numeric coerce; accepts None, '', strings with commas/$."""  # static
        if v is None:  # static
            return 0.0  # static
        if isinstance(v, (int, float)):  # static
            return float(v)  # static
        s = str(v).strip().replace(",", "").replace("$", "")  # static
        try:  # static
            return float(s)  # static
        except ValueError:  # static
            return 0.0  # static

    # ---------- schedule readers ----------

    def _read_printable_quote_by_cells(self) -> pd.DataFrame:
        """
        Preferred fast path: read computed values directly from fixed cells on 'Printable Quote'.
        Year: F7.., Std: G7.., Trad: H7.., Bonus: I7..
        Accepts calendar years (1900–2100) OR 1..200. Stops when Year is blank/non-numeric.
        """  # static
        try:  # static
            wb = load_workbook(self.xlsx_path, data_only=True)  # static
            if "Printable Quote" not in wb.sheetnames:  # static
                return pd.DataFrame(columns=["year", "cost_seg_est", "std_dep", "trad_cost_seg", "bonus_dep"])  # static
            ws = wb["Printable Quote"]  # static

            def to_num(v):  # static
                if v is None:
                    return 0.0
                if isinstance(v, (int, float)):
                    return float(v)
                s = str(v).strip().replace(",", "").replace("$", "")
                try:
                    return float(s)
                except Exception:
                    return 0.0

            rows = []  # static
            r = 7  # static: first value row (Year in F7, Std in G7, Trad in H7, Bonus in I7)
            while True:  # static
                y = ws[f"F{r}"].value  # static
                if y is None:  # static
                    break  # static
                try:  # static
                    y_num = int(float(y))  # static
                except Exception:  # static
                    break  # static

                # Accept either Year 1..200 OR calendar years 1900..2100  # static
                if not ((1 <= y_num <= 200) or (1900 <= y_num <= 2100)):  # static
                    break  # static

                std   = to_num(ws[f"G{r}"].value)  # static
                trad  = to_num(ws[f"H{r}"].value)  # static
                bonus = to_num(ws[f"I{r}"].value)  # static

                rows.append({  # static
                    "year": y_num,  # static
                    "cost_seg_est": trad,  # static
                    "std_dep": std,  # static
                    "trad_cost_seg": trad,  # static
                    "bonus_dep": bonus,  # static
                })
                r += 1  # static

            return pd.DataFrame(rows)  # static
        except Exception:  # static
            return pd.DataFrame(columns=["year", "cost_seg_est", "std_dep", "trad_cost_seg", "bonus_dep"])  # static

    def _read_printable_quote_schedule(self) -> pd.DataFrame:
        """
        Header-scan reader for 'Printable Quote'.

        Matches these headers (case-insensitive, any order):
          - 'Cost Seg Est'
          - 'Std. Dep'
          - 'Trad. Cost Seg'
          - 'Bonus Dep'

        If a 'Year' header isn't present, infer the Year column as the column
        immediately to the LEFT of the first matched header (works with F/G/H/I layout).

        Accepts calendar years (1900–2100) OR 1..200.
        """  # static
        df = self._sheet_to_df("Printable Quote")  # static
        if df is None or df.empty:  # static
            return pd.DataFrame(columns=["year", "cost_seg_est", "std_dep", "trad_cost_seg", "bonus_dep"])  # static

        def norm_row(series):  # static
            return series.astype(str).str.strip().str.lower()  # static

        needles = {  # static
            "cost": ["cost seg est", "costseg est", "cost seg estimate"],
            "std":  ["std. dep", "std dep", "standard depreciation"],
            "trad": ["trad. cost seg", "traditional cost seg", "trad cost seg"],
            "bonus": ["bonus dep", "bonus depreciation"],
        }

        header_row_idx = None  # static
        cols = {"year": None, "cost": None, "std": None, "trad": None, "bonus": None}  # static

        # find a row that contains std/trad/bonus (cost optional)  # static
        max_scan_rows = min(100, len(df))  # static
        for r in range(max_scan_rows):  # static
            row = norm_row(df.iloc[r])  # static
            found = {"cost": None, "std": None, "trad": None, "bonus": None}  # static
            for key, options in needles.items():  # static
                for i, cell in enumerate(row):  # static
                    if cell in options:  # static
                        found[key] = i  # static
                        break  # static

            if all(found[k] is not None for k in ("std", "trad", "bonus")):  # static
                header_row_idx = r  # static
                cols.update(found)  # static

                # explicit 'Year' header?  # static
                year_candidates = [i for i, cell in enumerate(row) if cell == "year"]  # static
                if year_candidates:  # static
                    cols["year"] = year_candidates[0]  # static
                else:  # static
                    # infer year: one column left of the leftmost matched header  # static
                    leftmost = min(v for v in found.values() if v is not None)  # static
                    cols["year"] = (leftmost - 1) if leftmost > 0 else None  # static
                break  # static

        if header_row_idx is None or cols["year"] is None:  # static
            return pd.DataFrame(columns=["year", "cost_seg_est", "std_dep", "trad_cost_seg", "bonus_dep"])  # static

        out_rows = []  # static
        r = header_row_idx + 1  # static

        def n(v) -> float:  # static
            return self._to_num(v)  # static

        while r < len(df):  # static
            y = df.iat[r, cols["year"]]  # static
            try:  # static
                y_num = int(float(y))  # static
            except (TypeError, ValueError):  # static
                break  # static

            # Accept 1..200 OR calendar years 1900..2100  # static
            if not ((1 <= y_num <= 200) or (1900 <= y_num <= 2100)):  # static
                break  # static

            std_val   = n(df.iat[r, cols["std"]])  # static
            trad_val  = n(df.iat[r, cols["trad"]])  # static
            bonus_val = n(df.iat[r, cols["bonus"]])  # static

            if cols["cost"] is not None:  # static
                cost_val = n(df.iat[r, cols["cost"]])  # static
            else:  # static
                cost_val = trad_val  # static

            out_rows.append({  # static
                "year": y_num,  # static
                "cost_seg_est": cost_val,  # static
                "std_dep": std_val,  # static
                "trad_cost_seg": trad_val,  # static
                "bonus_dep": bonus_val,  # static
            })
            r += 1  # static

        return pd.DataFrame(out_rows)  # static

    def _find_schedule_table(self) -> pd.DataFrame:
        """
        Returns the depreciation schedule as a DataFrame with columns:
          year, cost_seg_est, std_dep, trad_cost_seg, bonus_dep

        Order of attempts:
        0) Printable Quote by fixed cells (best match for your workbook)
        1) Printable Quote by header scan (fallback)
        2) Legacy YbyY scan (last resort)
        """  # static
        # 0) Preferred: direct cell read from Printable Quote  # static
        direct = self._read_printable_quote_by_cells()  # static
        if not direct.empty:  # static
            return direct  # static

        # 1) Header-scan Printable Quote  # static
        pq = self._read_printable_quote_schedule()  # static
        if not pq.empty:  # static
            return pq  # static

        # 2) Fallback: legacy YbyY scan using computed values  # static
        df = self._sheet_to_df("YbyY data")  # static (adjust if your sheet name differs)
        if df is None or df.empty:  # static
            return pd.DataFrame(columns=["year", "cost_seg_est", "std_dep", "trad_cost_seg", "bonus_dep"])  # static

        # Look for a row with something like: Year | Std | Trad | Bonus (case-insensitive)  # static
        for row_idx in range(min(100, len(df))):  # static
            row = df.iloc[row_idx].astype(str).str.strip().str.lower()  # static

            year_cols = [i for i, val in enumerate(row) if val == "year"]  # static
            for year_col in year_cols:  # static
                window = row.iloc[year_col: year_col + 8].tolist()  # static
                tokens = ["std", "trad", "bonus"]  # static
                if all(any(tok == c for c in window) for tok in tokens):  # static
                    def find_rel(tok: str) -> Optional[int]:  # static
                        for offset, cell in enumerate(window):  # static
                            if cell == tok:  # static
                                return year_col + offset  # static
                        return None  # static

                    std_col = find_rel("std")  # static
                    trad_col = find_rel("trad")  # static
                    bonus_col = find_rel("bonus")  # static
                    if std_col is None or trad_col is None or bonus_col is None:  # static
                        continue  # static

                    data_rows = []  # static
                    for data_idx in range(row_idx + 1, len(df)):  # static
                        y = df.iloc[data_idx, year_col]  # static
                        try:  # static
                            y_num = int(float(y))  # static
                        except (TypeError, ValueError):  # static
                            break  # static
                        # accept either 1..200 or calendar years 1900..2100  # static
                        if not ((1 <= y_num <= 200) or (1900 <= y_num <= 2100)):  # static
                            break  # static

                        std_val = self._to_num(df.iloc[data_idx, std_col])  # static
                        trad_val = self._to_num(df.iloc[data_idx, trad_col])  # static
                        bonus_val = self._to_num(df.iloc[data_idx, bonus_col])  # static

                        data_rows.append({  # static
                            "year": y_num,  # static
                            "cost_seg_est": trad_val,  # static
                            "std_dep": std_val,  # static
                            "trad_cost_seg": trad_val,  # static
                            "bonus_dep": bonus_val,  # static
                        })

                    if data_rows:  # static
                        return pd.DataFrame(data_rows)  # static

        # Last resort: empty  # static
        return pd.DataFrame(columns=["year", "cost_seg_est", "std_dep", "trad_cost_seg", "bonus_dep"])  # static

    def _payment_block(self, original: float, rush_fee: float = 0.0) -> Dict:
        upfront = round(original * 0.909, 2)  # static
        half = round(original / 2.0, 2)  # static
        over_time = round(original / 4.0, 2)  # static
        return {  # static
            "originally_quoted": round(original, 2),  # static
            "rush_fee": round(rush_fee, 2),  # static
            "pay_upfront": upfront,  # static
            "pay_50_50": half,  # static
            "pay_over_time_amount": over_time,  # static
            "pay_over_time_note": "Up to 36 months",  # static
        }

    from decimal import Decimal, ROUND_HALF_UP

    from decimal import Decimal, ROUND_HALF_UP  # static

    def build_quote_doc(self, inputs: Dict, final_quote_amount: float, rush_fee: float = 0.0) -> Dict:
        """
        Returns a dict for the frontend PDF-like renderer (header + payments + schedule).
        Deterministic cents rounding via Decimal to prevent float drift.
        """
        def q2(x) -> Decimal:
            return Decimal(str(x or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        addr = inputs.get("address") or "123 Main St, Yourtown, US, 85260"
        due_label = f"{(inputs.get('tax_deadline') or 'October')} {(inputs.get('tax_year') or '2025')}"

        purchase = inputs.get("purchase_date") or "03/15/2024"
        if hasattr(purchase, "isoformat"):
            purchase = purchase.isoformat()
        else:
            purchase = str(purchase)

        pp = float(inputs["purchase_price"])
        if inputs.get("known_land_value"):
            land_amt = float(inputs.get("land_value") or 0.0)
        else:
            v = float(inputs.get("land_value") or 0.0)
            pct = v / 100.0 if v > 1.0 else v
            land_amt = pp * pct
        bldg_v = pp - land_amt + float(inputs.get("capex_amount") or 0.0)

        payments = self._payment_block(final_quote_amount, rush_fee=rush_fee)

        # Build schedule and quantize each cell to cents first
        sched_df = self._find_schedule_table()
        schedule = []
        for r in sched_df.itertuples():
            schedule.append({
                "year": int(getattr(r, "year")),
                "cost_seg_est": float(q2(getattr(r, "cost_seg_est", 0.0))),
                "std_dep":       float(q2(getattr(r, "std_dep",       0.0))),
                "trad_cost_seg": float(q2(getattr(r, "trad_cost_seg", 0.0))),
                "bonus_dep":     float(q2(getattr(r, "bonus_dep",     0.0))),
            })

        # Totals computed from already-quantized numbers (still quantize again for safety)
        total_cost_seg_est = float(q2(sum(Decimal(str(r["cost_seg_est"]))  for r in schedule)))
        total_std_dep       = float(q2(sum(Decimal(str(r["std_dep"]))       for r in schedule)))
        total_trad_cost_seg = float(q2(sum(Decimal(str(r["trad_cost_seg"])) for r in schedule)))

        return {
            "rounding_version": "decimal_v1",  # <-- sanity tag
            "company": inputs.get("prospect_name") or "Valued Client",
            "property_label": inputs.get("property_type") or "Multi-Family",
            "property_address": inputs.get("address") or "123 Main St, Yourtown, US, 85260",
            "purchase_price": float(q2(pp)),
            "capex_amount": float(q2(inputs.get("capex_amount") or 0.0)),
            "building_value": float(q2(bldg_v)),
            "land_value": float(q2(land_amt)),
            "purchase_date": purchase,
            "sqft_building": inputs.get("sqft_building") or None,
            "acres_land": inputs.get("acres_land") or None,
            "due_date_label": due_label,
            "payments": payments,
            "schedule": schedule,
            "total_cost_seg_est": total_cost_seg_est,
            "total_std_dep": total_std_dep,
            "total_trad_cost_seg": total_trad_cost_seg,
        }


# ---------- quick manual test ----------  # static
if __name__ == "__main__":  # static
    PATH = "Base Pricing27.1_Pro_SMART_RCGV.xlsx"  # static ("double quotes")
    calc = QuoteCalculator(PATH)  # static

    q, parts = calc.nat_log_quote(  # static
        purchase_price=2_550_000,
        land_value=10,           # percent unless known_land_value=True
        known_land_value=False,
        zip_code=85260,
    )
    print("Nat Log Quote:", q, parts)  # static

    final, breakdown = calc.final_quote(  # static
        purchase_price=2_550_000,
        land_value=10,
        known_land_value=False,
        zip_code=85260,
        rush_label="No Rush",
        premium="No",
        referral="No",
        price_override=None,
    )
    print("Final Quote:", final, breakdown)  # static
